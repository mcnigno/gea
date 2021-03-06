from app import db
from flask_appbuilder.models.sqla.interface import SQLAInterface
from flask import flash, send_file, make_response, redirect, url_for

from .models import (Matrix, Document,Job, Discipline, Unit, Application, Doctype, Subdoctype, Domain, Partner,
                     Cdrlitem, Documentclass, Mr, Vendor, DocRequests)
#from .views import send_csv
import csv, xlsxwriter
from werkzeug.utils import secure_filename
import uuid
import openpyxl, os 
from app import app





def adddoc3(self, item):
    # Set the Request type
    if item.vendor and item.mr:
        item.request_type = 'vendor'
    
    session = db.session
    matrix = session.query(Matrix)
    if item.unit.unit == '000':
        item_matrix = str.join('-', (item.unit.unit,
                                     item.materialclass.materialclass,
                                     item.doctype.doctype,
                                     # item.sheet,
                                     item.partner.partner))
    else:
        item_matrix = str.join('-', (item.unit.unit,
                                     item.materialclass.materialclass,
                                     item.doctype.doctype,
                                     # item.sheet,
                                     ))

    item_serial = str.join('-', (item.unit.unit,
                                 item.materialclass.materialclass,
                                 item.doctype.doctype,
                                 # item.sheet,
                                 ))

    print('controllo item_ matrix', item_matrix)
    found = False
    for row in matrix:
        print('loop controllo matrix uguali',
              row.matrix, item_matrix,
              item.matrix)
        if row.matrix == item_matrix:
            print('trovate matrix uguali')
            print('row counter prima:', row.counter)
            row.counter += 1
            print('row counter dopo:', row.counter)
            datamodel = SQLAInterface(Matrix, session=session)
            datamodel.edit(row)

            item.matrix_id = row.id
            code = item_serial + "-" + str(row.counter).zfill(5) + "-" + item.sheet

            datamodel = SQLAInterface(Document, session=session)
            doc = Document(docrequests_id=item.id, code=code)
            datamodel.add(doc)

            message = 'Your code is ' + code
            flash(message, category='info')
            found = True
    
    # Matrix Not Found
    if found is False:
        print('Matrix NOT FOUND')

        if str(item.unit) == '000':
            print('found unit 000')
            jv = {
                    'TTSJV': 50000,
                    'TPIT': 60000
                }

            # Add New Matrix
            datamodel = SQLAInterface(Matrix, session=session)

            print('counter', jv['TTSJV'], item.partner, jv[str(item.partner)])
            print('item matrix:', item_matrix)

            matrix = Matrix(counter=jv[str(item.partner)] + 1,
                            matrix=item_matrix)

            print('-----2------')

            datamodel.add(matrix)

            # Add New Doc with quantity jv + 1
            datamodel = SQLAInterface(Document, session=session)
            code = item_serial + "-" + str(jv[str(item.partner)] + 1).zfill(5) + "-" + item.sheet
            doc = Document(docrequests_id=item.id, code=code)

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')

        else:
            # Add New Matrix
            datamodel = SQLAInterface(Matrix, session=session)
            print('item matrix:', item_matrix)

            print('item matrix after:', item_matrix)
            matrix = Matrix(matrix=item_matrix)
            datamodel.add(matrix)

            # Add New Doc with quantity 1
            datamodel = SQLAInterface(Document, session=session)
            code = item_serial + "-" + "1".zfill(5) + "-" + item.sheet
            doc = Document(docrequests_id=item.id, code=code)

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')

    db.session.flush()

def bapco(self, item):
    # Set the DB session
    session = db.session

    # Set tehe Document item
    print('document setting', item.job_id, 
                            item.discipline_id, 
                            item.unit_id, 
                            item.application_id, 
                            item.doctype_id, 
                            item.subdoctype_id, 
                            item.domain_id
                            )

    doc = Document(job_id=item.job_id,discipline_id=item.discipline_id, unit_id=item.unit_id, application_id=item.application_id, domain_id=item.domain_id,
                   doctype_id=item.doctype_id, subdoctype_id=item.subdoctype_id, partner_id=item.partner_id)
    
    print('document setted')
    if item.documentclass:
        print('documentclass', item.documentclass, item.documentclass_id )
        doc.documentclass_id = item.documentclass_id
    
    if item.cdrlitem:
        print('cdrlitem', item.cdrlitem, item.cdrlitem_id )
        doc.cdrlitem_id = item.cdrlitem_id

    # Set the Request type
    if item.vendor and item.mr:
        item.request_type = 'vendor'
        doc.vendor_id = item.vendor_id
        doc.mr_id = item.mr_id

    else:
        item.request_type = 'engineering'
    
    # Set item_matrix based on unit type
    print(item.unit.unit)
    result = db.session.query(Unit).filter(Unit.unit == str(item.unit.unit)).first()
    print("result", result)

    if str(result.unit_type) == 'common':
        print('Match unit type common Found')

        # Add the partner id to the matrix
        item_matrix = str.join('-', (
                                     str(item.job.job),
                                     str(item.discipline.discipline),
                                     str(item.unit.unit)+ str(item.application.application),
                                     
                                     str(item.doctype.doctype) + str(item.subdoctype.subdoctype),
                                     str(item.domain.domain),
                                     # item.sheet,
                                     str(item.partner)
                                     ))
    
    else:

        item_matrix = str.join('-', (
                                     str(item.job.job),
                                     str(item.discipline.discipline),
                                     str(item.unit.unit) + str(item.application.application), 
                                     
                                     str(item.doctype.doctype) + str(item.subdoctype.subdoctype),
                                     str(item.domain.domain)
                                     ))
    # Set the bapco base code
    item_serial = str.join('-', (
                                str(item.job.job),
                                str(item.discipline.discipline),
                                str(item.unit.unit) + str(item.application.application),
                                
                                str(item.doctype.doctype) + str(item.subdoctype.subdoctype),
                                str(item.domain.domain)
                                 # item.sheet,
                                 ))
    
    # Increment the Matrix counter or create a new one
   
    
    matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()
    partner = db.session.query(Partner).filter(Partner.partner == str(item.partner)).first()
    print('item_matrix, matrix',item_matrix, matrix)

    if matrix:
        print('matrix, counter',matrix, matrix.counter)
        
        if matrix.counter + 1 <= result.stop or matrix.counter + 1 <= partner.common_stop:
            print('matrix, counter',matrix, matrix.counter)
            #
            print('you are here')
            matrix.counter += 1
            datamodel = SQLAInterface(Matrix, session=session)
            datamodel.edit(matrix)

            item.matrix_id = matrix.id
            doc.matrix_id = matrix.id
            code = item_serial + str(matrix.counter).zfill(3)

            datamodel = SQLAInterface(Document, session=session)
            
            doc.docrequests_id = item.id
            doc.code = code
        
            datamodel.add(doc)

            message = 'Your code is ' + code
            flash(message, category='info')
        else:
            flash('No more Numbers available for this combination.', category='warning')
        
            
    else:
        # Create a New Matrix for common units
        if result.unit_type == 'common':

            print('item partner to find: ', item.partner)
            
            partner = db.session.query(Partner).filter(Partner.partner == str(item.partner.partner)).first()
           
            matrix = Matrix(counter=partner.common_start + 1, matrix=str(item_matrix))
            datamodel = SQLAInterface(Matrix, session=session)
            datamodel.add(matrix)

            matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()


            # Add new Doc with quantity partner common start + 1
            datamodel = SQLAInterface(Document, session=session)
            
            code = item_serial + str(partner.common_start + 1).zfill(3)
            
            #doc = Document(docrequests_id=item.id, code=code)
            doc.matrix_id = matrix.id
            doc.docrequests_id = item.id
            doc.code = code

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')
        
        else:
            # Create a new Matrix for standard units
            
            datamodel = SQLAInterface(Matrix, session=session)
            matrix = Matrix(counter=result.start + 1, matrix=item_matrix)
            datamodel.add(matrix)  

            matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()
        

            # Add new Doc with quantity 1
            datamodel = SQLAInterface(Document, session=session)
            code = item_serial + str(result.start + 1).zfill(3) 
            #doc = Document(docrequests_id=item.id, code=code)
            doc.matrix_id = matrix.id
            doc.docrequests_id = item.id
            doc.code = code

            datamodel.add(doc)
            message = 'Your code is ' + code
            flash(message, category='info')

    db.session.flush()
    return code

def init_bapco(self, item):
    # Set the DB session
    session = db.session

    id_unit = db.session.query(Unit).filter(Unit.id == str(item.unit_id)).first()
    print('id_unit ',id_unit)
    id_materialclass = db.session.query(Materialclass).filter(Materialclass.id == str(item.materialclass_id)).first()
    id_doctype = db.session.query(Doctype).filter(Doctype.id == str(item.doctype_id)).first()
    id_patner = db.session.query(Partner).filter(Partner.id == str(item.partner_id)).first()
    id_cdrlitem = None
    id_documentclass = None
    id_vendor = None
    id_mr = None
    
    req = DocRequests(unit_id=id_unit.id, materialclass_id=id_materialclass.id, 
                      doctype_id=id_doctype.id, partner_id=id_patner.id,
                      request_type=item.request_type)
    
    doc = Document(unit_id=id_unit.id, materialclass_id=id_materialclass.id,
                   doctype_id=id_doctype.id, partner_id=id_patner.id)

    if item.cdrlitem_id:
        id_cdrlitem = db.session.query(Cdrlitem).filter(Cdrlitem.id == str(item.cdrlitem_id)).first()
        req.cdrlitem_id = id_cdrlitem.id
        doc.cdrlitem_id = id_cdrlitem.id
        
    if item.documentclass_id:
        id_documentclass = db.session.query(Documentclass).filter(Documentclass.id == str(item.documentclass_id)).first()
        req.documentclass_id = id_documentclass.id
        doc.documentclass_id = id_documentclass.id

    # Set the Request type
    if item.vendor and item.mr:
        req.request_type = 'vendor'
        id_vendor = db.session.query(Vendor).filter(Vendor.id == str(item.vendor_id)).first()
        id_mr = db.session.query(Mr).filter(Mr.id == str(item.mr_id)).first()
        req.vendor_id = id_vendor.id
        req.mr_id = id_mr.id

        doc.vendor_id = id_vendor.id
        doc.mr_id = id_mr.id
    
    else:
        req.request_type = 'engineering'
    
    # Set item_matrix based on unit type
    result = db.session.query(Unit).filter(Unit.id == str(item.unit_id)).first()
    
    if str(result.unit_type) == 'common':
        print('Match unit type common Found')

        # Add the partner id to the matrix
        item_matrix = str.join('-', (str(id_unit.unit),
                                     str(id_materialclass.materialclass),
                                     str(id_doctype.doctype),
                                     # item.sheet,
                                     str(id_patner.partner)
                                     ))
    else:
        item_matrix = str.join('-', (str(id_unit.unit),
                                     str(id_materialclass.materialclass),
                                     str(id_doctype.doctype),
                                     # item.sheet,
                                     ))
    # Set the bapco base code
    item_serial = str.join('-', (str(id_unit.unit),
                                 str(id_materialclass.materialclass),
                                 str(id_doctype.doctype),
                                 # item.sheet,
                                 ))
    
    # Increment the Matrix counter or create a new one
    print('Matrix to search:', 
    (str(id_unit.unit),
                                     str(id_materialclass.materialclass),
                                     str(id_doctype.doctype),
                                     # item.sheet,
                                     str(id_patner.partner)
                                     )
    )
    matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()
    if matrix:
        matrix.counter += 1
        datamodel = SQLAInterface(Matrix, session=session)
        datamodel.edit(matrix)

        req.matrix_id = matrix.id
        code = item_serial + "-" + str(matrix.counter).zfill(5) + "-" + item.sheet
        
        datamodel = SQLAInterface(Document, session=session)
        
        #doc = Document(docrequests=req, code=code)
        doc.docrequests = req
        doc.code = code
        datamodel.add(doc)

        message = 'Your code is ' + code
        flash(message, category='info')
    else:
        # Create a New Matrix for common units
        if result.unit_type == 'common':

            print('item partner to find: ', id_patner.partner)
            
            partner = db.session.query(Partner).filter(Partner.id == str(item.partner_id)).first()
           
            matrix = Matrix(counter=partner.common_start + 1, matrix=str(item_matrix))
            datamodel = SQLAInterface(Matrix, session=session)
            datamodel.add(matrix)
            
            # Find the Matrix'ID
            id_matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()
            req.matrix_id = id_matrix.id 

            # Add new Doc with quantity partner common start + 1
            
            code = item_serial + "-" + str(partner.common_start + 1).zfill(5) + "-" + item.sheet
            
            datamodel = SQLAInterface(Document, session=session)
            
            #doc = Document(docrequests=req, code=code)
            doc.docrequests = req
            doc.code = code
            
            datamodel.add(doc)
            
            message = 'Your code is ' + code
            flash(message, category='info')
        else:
            # Create a new Matrix for standard units
            datamodel = SQLAInterface(Matrix, session=session)
            matrix = Matrix(matrix=item_matrix)
            datamodel.add(matrix)
            
            # Find the Matrix'ID
            id_matrix = db.session.query(Matrix).filter(Matrix.matrix == item_matrix).first()
            req.matrix_id = id_matrix.id

            # Add new Doc with quantity 1
            
            code = item_serial + "-" + "1".zfill(5) + "-" + item.sheet
            
            datamodel = SQLAInterface(Document, session=session)
            
            #doc = Document(docrequests=req, code=code)   
            doc.docrequests = req
            doc.code = code

            datamodel.add(doc)
            
            message = 'Your code is ' + code
            flash(message, category='info')

    db.session.flush()
    return code

def tocsv(self, item,  codes_list):
    print('tocsv FUNCTION')
    print(codes_list)
    filename = 'app/static/csv/bapco_request_'+ str(item.id) + '.csv'
    with open(filename, 'w') as csv_file:
        # writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
        writer = csv.writer(csv_file, dialect='excel')
        head = [['Id Code', 'Bapco Code', 'Contractor Code']] 
        writer.writerows(head)
        writer.writerows(codes_list)
        print('inside the with file for CV rendering')
        #print(csv_file)
    
    print('BEFORE file SEND')
    #send_file(filename, as_attachment=True)
    print('file SEND')
    '''
    response = make_response(filename)
    cd = 'attachment; filename=mycsv.csv'
    response.headers['Content-Disposition'] = cd
    response.mimetype = 'text/csv'
    return response
    '''
def toxlsx(self, item,  codes_list):
    print('toXLSX FUNCTION')
    print('this is the code list', codes_list)
    #filename = 'app/static/csv/bapco_request_'+ str(item.id) + '.xlsx'

    workbook = xlsxwriter.Workbook('app/static/csv/bapco_request_'+ str(item.id) + '.xlsx')
    workbook.set_properties({
                            'title':    'Bapco Request spreadsheet',
                            'subject':  'With document properties',
                            'author':   'Quasar DCC Team',
                            'manager':  'Danilo Pacifico',
                            'company':  'Quasar',
                            'category': 'Bapco spreadsheets',
                            'keywords': 'Bapco, Bapco Codes, Properties',
                            'comments': 'Created by webtools.quasarPM.com'})

    worksheet = workbook.add_worksheet(name='Bapco Request List')

    bold = workbook.add_format({'bold': True})
    worksheet.set_column('A:A', 25)
    worksheet.set_column('B:B', 25)
    worksheet.write('A1', 'Bapco Code', bold)
    worksheet.write('B1', 'Contractor Code', bold)
    # start after the header
    row = 1
    col = 0
    print('this is the touple')
    print(tuple(codes_list))
    #worksheet.write(0,0,'something')
    t_list = tuple(codes_list)
    for code in (t_list):
        #print('Looping colist', code, row)
        worksheet.write(row, col, str(code[0]))
        worksheet.write(row, col)
        row += 1
    workbook.close()

def codes_to_xlsx(codes_list):
    print('CODES toXLSX FUNCTION')
    print(codes_list)
    #filename = 'app/static/csv/bapco_request_'+ str(item.id) + '.xlsx'
    filename = 'Quasar|' + str(uuid.uuid4()) + '|bapco.xlsx'
    workbook = xlsxwriter.Workbook('app/static/csv/' + filename)
    workbook.set_properties({
                            'title':    'Bapco Request spreadsheet',
                            'subject':  'With document properties',
                            'author':   'Quasar DCC Team',
                            'manager':  'Danilo Pacifico',
                            'company':  'Quasar',
                            'category': 'Bapco spreadsheets',
                            'keywords': 'Bapco, Bapco Codes, Properties',
                            'comments': 'Created by webtools.quasarPM.com'})

    worksheet = workbook.add_worksheet(name='Bapco Request List')

    bold = workbook.add_format({'bold': True})
    worksheet.set_column('A:A', 25)
    worksheet.set_column('B:B', 25)
    worksheet.write('A1', 'Bapco Code', bold)
    worksheet.write('B1', 'Contractor Code', bold)
    # start after the header
    row = 1
    col = 0
    print(tuple(codes_list))
    #worksheet.write(0,0,'something')
    #t_list = tuple(codes_list)
    t_list = codes_list
    
    #print('*********', t_list)
    if isinstance(t_list, list):
        t_list = tuple(codes_list)
    else:
        t_list = codes_list
    for code in (t_list):
        print('Looping colist', code, row)
        worksheet.write(row, col, str(code[0]))
        worksheet.write(row, col+1, str(code[1]))
        row += 1
    workbook.close()
    return filename

def update_from_xlsx(file):
    session = db.session
    print('update FUNCTION!')
    book = openpyxl.load_workbook(file)
    sheet = book.active
    a1 = sheet['A1']
    print(a1.value)
    reserved_list = []
    updated_list = []
    for row in sheet.iter_rows(min_row=2):
        bapco_code = row[0].value
        oldcode = row[1].value
        print(bapco_code, oldcode)
        doc = db.session.query(Document).filter(Document.code == str(bapco_code)).first()

        if doc and doc.oldcode == 'empty':
            print('this is the ID' ,doc.id)
            datamodel = SQLAInterface(Document, session=session)
            print('BEFORE oldcode', doc.oldcode)
            doc.oldcode = oldcode
            updated_list.append([doc.id, doc.code, doc.oldcode])
            datamodel.edit(doc)

        else:
            reserved_list.append([doc.id, doc.code, doc.oldcode])
    return reserved_list, updated_list


def setting_update(file):
    session = db.session
    print('Setting Update start')
    book = openpyxl.load_workbook(file)
    sheet = book.active
    set_class = sheet['A1'].value
    print('the setting class:', set_class)

    reserved_list = []
    updated_list = []

    for row in sheet.iter_rows(min_row=2):
        param = row[0].value
        name = row[1].value
        desc = row[2].value
        print(param, name, desc)
        set_dict = {
            'Job': [Job, Job.job],
            'Discipline': [Discipline, Discipline.discipline],
            'Unit': [Unit, Unit.unit],
            'Application': [Application, Application.application],
            'Doctype': [Doctype, Doctype.doctype],
            'Subdoctype': [Subdoctype, Subdoctype.subdoctype],
            'Domain': [Domain, Domain.domain],
            'Cdrlitem': [Cdrlitem, Cdrlitem.cdrlitem],
            'Documentclass': [Documentclass, Documentclass.documentclass],
            'Mr': [Mr, Mr.mr],
            'Vendor': [Vendor, Vendor.vendor],
            'Partner': [Partner, Partner.partner]
        }
        # Update the setting if a param already exist
        tmp_class = set_dict[set_class][0]
        tmp_param = set_dict[set_class][1]
        my_class = db.session.query(tmp_class).filter(tmp_param == str(param)).first()
        
        datamodel = SQLAInterface(tmp_class, session=session)
    
        if my_class:
            print(my_class)
            my_class.name = name
            my_class.description = desc
            datamodel.edit(my_class)
            updated_list.append([my_class.id, my_class.name, my_class.description])
        else:
            # or Create new record in setting
            
            my_class = tmp_class()
            if set_class == 'Unit':
                my_class.unit = param
            elif set_class == 'Job':
                my_class.job = param
            elif set_class == 'Discipline':
                my_class.discipline = param
            elif set_class == 'Application':
                my_class.application = param
            elif set_class == 'Doctype':
                my_class.doctype = param
            elif set_class == 'Subdoctype':
                my_class.subdoctype = param
            elif set_class == 'Domain':
                my_class.domain = param
            elif set_class == 'Cdrlitem':
                my_class.cdrlitem = param
            elif set_class == 'Documentclass':
                my_class.documentclass = param
            elif set_class == 'Mr':
                my_class.mr = param
            elif set_class == 'Vendor':
                my_class.vendor = param
            elif set_class == 'Partner':
                my_class.partner = param
            
            else:
                return print('Setting Class NOT Found')
             
            my_class.name = name
            my_class.description = desc
            datamodel.add(my_class)
            reserved_list.append([my_class.id, my_class.name, my_class.description])

    return reserved_list, updated_list

def old_codes_update(self, file):
    book = openpyxl.load_workbook(file)
    sheet = book.active
    #header = sheet['A1':'M1']
    '''
    unit = header[0][0].value
    materialclass = header[0][1].value
    doctype = header[0][2].value
    bapco_code = header[0][5].value
    description = header[0][6].value
    oldcode = header[0][7].value
    cdrlitem = header[0][8].value
    documentclass =header[0][9].value
    note = header[0][10].value
    trasmittal = header[0][11].value
    '''
    session = db.session
    datamodel = SQLAInterface(DocRequests, session=session)

    for row in sheet.iter_rows(min_row=2):
        req = DocRequests()
        req.quantity = 1
        req.unit = Unit(unit=row[0].value) 
        req.materialclass = Materialclass(materialclass=row[1].value) 
        req.doctype = Doctype(doctype=row[2].value) 
        req.cdrlitem = Cdrlitem(cdrlitem=row[8].value) 
        req.documentclass = Documentclass(documentclass=row[9].value) 
        req.partner = Partner(partner=row[12].value) 
        #req.oldcode = row[7].value
        print('search for Unit:', req.unit.unit)
        #code = bapco(self, req)
        #row[14] = code
        datamodel.add(req)
        
    reserved_list = ['coming soon...']
    updated_list = ['coming soon...']
    book.close()
    return reserved_list, updated_list

def gen_excel_byreq(req_item):
    session = db.session
    req = session.query(Document).filter(Document.docrequests_id == req_item.id).all()

    codes_list = []
    for item in req:
        codes_list.append([item.code,item.oldcode])
    
    filename = codes_to_xlsx(codes_list)
    return filename
    
#xls = open('bapco_codes.xlsx','rb')
def old_codes(self, file):

    book = openpyxl.load_workbook(file)
    sheet = book.active
    session = db.session
    # Create the datamodel
    #datamodel = SQLAInterface(DocRequests, session=session)
    # first_req = DocRequests(unit=Unit(unit='QSR'), materialclass=Materialclass(materialclass='1'), doctype=Doctype(doctype='QSR'), partner=Partner(partner='QSR'))
    #datamodel.add(first_req)
    

    
    found_list = []
    not_found_list = []
    for row in sheet.iter_rows(min_row=2):
        check = True
        datamodel = SQLAInterface(DocRequests, session=session)
        req = DocRequests()
        #req.id = 1
       
        # Unit Query
        if row[0].value:
            print('row[0] value is:' ,row[0].value)
            try:
                unit_id = session.query(Unit).filter(Unit.unit == row[0].value).first()
                print('query: unit_id is:', unit_id)
                print('query: unit_id.id is:', unit_id.id)
                print('query: unit_id.unit is:', unit_id.unit)
                print('query: unit_id.unit_type is:', unit_id.unit_type)
                
                req.unit_id = unit_id.id
                print('stop here')
                #req.unit = unit_id.unit
                
                #req.unit.unit = unit_id.unit
                print('stop 2 here')
                '''
                if unit_id.unit_type == '000':
                    print('checking common')
                    #req.unit_type = 'common'
                '''
                print('req_unit_id is : ', req.unit_id)
                found_list.append(['Unit', req.unit_id, row[0].value])
            except:
                check = False

                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Unit Not Found',row[0].value ])
        print('after unit query the check is: ', check)
        # Materialclass Query
        if row[1].value:
            try:
                mat_id = session.query(Materialclass).filter(Materialclass.materialclass == row[1].value).first()
                req.materialclass_id = mat_id.id
                #req.materialclass = mat_id.materialclass
                found_list.append(['Materialclass', req.materialclass_id, row[1].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Material class Not Found',row[1].value ])
        
        # Doctype Query
        if row[2].value:
            try:
                doc_id = session.query(Doctype).filter(Doctype.doctype == row[2].value).first()
                req.doctype_id = doc_id.id
                #req.doctype = doc_id.doctype 
                found_list.append(['Doctype', req.doctype_id, row[2].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Doctype Not Found',row[2].value ])

        # CdrlItem Query
        print('cdrl item before check:', row[8].value)
        if row[8].value:
            print('cdrl item',row[8])
            try:
                cdrl_id = session.query(Cdrlitem).filter(Cdrlitem.cdrlitem == str(row[8].value)).first()
                req.cdrlitem_id = cdrl_id.id
                print('we got the crdrl item', req.cdrlitem_id)
                #req.cdrlitem = cdrl_id.cdrlitem
                found_list.append(['Cdrlitem', req.cdrlitem_id, row[8].value])
            except:
                #check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Cdrl Not Found',row[8].value ])
        
        # Documentclass Query
        if row[9].value:
            print('document class',row[9])
            try:
                dc_id = session.query(Documentclass).filter(Documentclass.documentclass == str(row[9].value)).first()
                req.documentclass_id = dc_id.id
                #req.documentclass = dc_id.documentclass
                found_list.append(['Documentclass', req.documentclass_id, row[9].value])
            except:
                #check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Document class Not Found',row[9].value ])

        # Partner Query
        if row[12].value:
            try:
                
                pa_id = session.query(Partner).filter(Partner.partner == row[12].value).first()
                req.partner_id = pa_id.id
                #req.partner = pa_id.partner
                found_list.append(['Partner', req.partner_id, row[12].value])
            except:
                check = False
                not_found_list.append(['bapco: '+row[5].value+' for '+row[7].value,'Partner Not Found', row[12].value ])
        
        # Add Sheet 001
        req.sheet = '001'
        #
        # 
        # ADD the first Request with id = 1
        #
        '''
        datamodel = SQLAInterface(DocRequests, session=session)
        first_req = DocRequests(unit='001', materialclass='A', doctype='HDD', partner='QSR')
        datamodel.add(first_req)
        req.id = 1
        '''
        
        if check is True:
            code = init_bapco(self, req)
            ask_bapco_code = row[13].value
            row[13].value = code
            
            
            print('Added DocRequstest:', req.unit_id, req.materialclass_id, req.doctype_id,
                    req.cdrlitem_id, req.documentclass_id, req.partner_id)
            print('Ask Bapco:', code)
            print(' Your Bapco Code: ',row[5].value)
            print('Your Code: ',row[7].value )

            
        else:
            print('Not Added DocRequstest:', req.unit_id, req.materialclass_id, req.doctype_id,
                    req.cdrlitem_id, req.documentclass_id, req.partner_id)
            print('Wrong Request: ',row[0].value,row[1].value,row[2].value,row[8].value,row[9].value,row[12].value )
    
    result_file = 'app/static/csv/upload_results.xlsx'
    book.save(result_file)


    '''
    print('Found List')
    for i in found_list:
        print(i[0],i[1], i[2])
    
    print('NOT Found List')
    for i in not_found_list:
        print(i[0],i[1], i[2])
    '''
    return not_found_list, found_list, result_file

#
# Mail Support
#


app.config['MAIL_SERVER']='mail.quasarpm.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USERNAME'] = 'info@quasarpm.com'
app.config['MAIL_PASSWORD'] = '300777Info'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
#app.config['MAIL_ASCII_ATTACHMENTS']= True

from flask_mail import Mail, Message

mail = Mail(app)
#user = get_user()
def mailsupport(codes, filename):
    msg = Message('GEA | DCC Support', sender = 'info@quasarpm.com', recipients = ['danilo.pacifico@gmail.com'])
    list_codes = []
    for row in codes:
        list_codes.append(' | '.join(row))  
    body = '<br>'.join(list_codes)  
    msg.body = body + "\n"
    msg.html = "<b>Ask Bapco | Support Request</b>\
    <p>" + body + "\
    <p>We have received your request and will contact you via email or by telephone shortly. If you require immediate assistance, please call our office. </p>\
    <p>If you do not receive an email response from us in your inbox within a day or so, please check your spam/junk email folder.</p>"
    mail.send(msg)

    return "Sent"
